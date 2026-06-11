#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BarTender 标签打印工具 v2.1
功能：集成 BarTender 2022 R2 Enterprise，IMEI 标签打印，Excel 数据校验
"""

import os
import sys
import csv
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import threading

# BarTender COM 接口
try:
    import win32com.client
    import pythoncom
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

# Excel 支持
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class PrintRecord:
    """打印记录"""
    def __init__(self, imei, print_time, copies, status="成功"):
        self.imei = imei
        self.print_time = print_time
        self.copies = copies
        self.status = status

    def to_dict(self):
        return {
            "imei": self.imei,
            "print_time": self.print_time,
            "copies": self.copies,
            "status": self.status
        }


class BarTenderPrintApp:
    """BarTender 标签打印应用"""
    
    VERSION = "v2.2.3"
    
    def __init__(self):
        # 先创建主窗口
        self.root = tk.Tk()
        self.root.title(f"BarTender 标签打印工具 {self.VERSION}")
        self.root.geometry("950x750")
        self.root.minsize(850, 700)
        
        # 配置文件路径
        self.app_dir = os.path.join(os.path.expanduser("~"), ".bartender-printer")
        os.makedirs(self.app_dir, exist_ok=True)
        self.config_file = os.path.join(self.app_dir, "bt_config.json")
        self.records_file = os.path.join(self.app_dir, "print_records.csv")
        
        # BarTender 相关
        self.bt_app = None
        self.bt_format = None
        
        # 数据
        self.print_records = []
        self.excel_data = []
        self.excel_file_path = ""
        
        # 初始化变量（必须在 tk.Tk() 之后）
        self.init_vars()
        
        # 加载配置和记录
        missing_files = self.load_config()
        self.load_records()
        
        # 创建 UI
        self.create_ui()
        
        # 刷新历史和统计
        self.refresh_history()
        self.refresh_stats()
        
        # 弹窗提示缺失的文件
        if missing_files:
            self.root.after(500, lambda: messagebox.showwarning(
                "配置文件缺失",
                "以下配置的文件未找到：\n\n" + "\n".join(missing_files) + "\n\n请重新选择这些文件。"
            ))
        
        # 绑定关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def init_vars(self):
        """初始化变量"""
        self.template_path_var = tk.StringVar()
        self.datasource_var = tk.StringVar(value="IMEI1")
        self.excel_path_var = tk.StringVar()
        self.excel_column_var = tk.StringVar(value="IMEI1")
        self.printer_var = tk.StringVar()
        self.copies_var = tk.IntVar(value=1)
        self.verify_excel_var = tk.BooleanVar(value=True)
    
    def create_ui(self):
        """创建用户界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 顶部标题
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(title_frame, text="BarTender 标签打印工具", font=("微软雅黑", 16, "bold")).pack(side=tk.LEFT)
        ttk.Button(title_frame, text="设置", command=self.open_settings).pack(side=tk.RIGHT)
        
        # 创建选项卡
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # 选项卡1：打印
        self.create_print_tab()
        
        # 选项卡2：历史记录
        self.create_history_tab()
        
        # 选项卡3：统计
        self.create_stats_tab()
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(fill=tk.X, pady=(10, 0))
        
        # 初始化 BarTender
        self.init_bartender()
    
    def create_print_tab(self):
        """创建打印选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="打印")
        
        # BarTender 模板配置
        template_frame = ttk.LabelFrame(tab, text="BarTender 模板", padding="10")
        template_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 模板文件
        file_frame = ttk.Frame(template_frame)
        file_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(file_frame, text="模板文件：", width=12).pack(side=tk.LEFT)
        self.template_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.template_path_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(file_frame, text="浏览", command=self.browse_template).pack(side=tk.RIGHT)
        
        # 数据源名称
        datasource_frame = ttk.Frame(template_frame)
        datasource_frame.pack(fill=tk.X)
        
        ttk.Label(datasource_frame, text="数据源名称：", width=12).pack(side=tk.LEFT)
        self.datasource_var = tk.StringVar(value="IMEI1")
        ttk.Entry(datasource_frame, textvariable=self.datasource_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Excel 数据文件配置
        excel_frame = ttk.LabelFrame(tab, text="IMEI 数据源（Excel）", padding="10")
        excel_frame.pack(fill=tk.X, pady=(0, 10))
        
        excel_file_frame = ttk.Frame(excel_frame)
        excel_file_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(excel_file_frame, text="Excel 文件：", width=12).pack(side=tk.LEFT)
        self.excel_path_var = tk.StringVar()
        ttk.Entry(excel_file_frame, textvariable=self.excel_path_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(excel_file_frame, text="选择文件", command=self.browse_excel).pack(side=tk.RIGHT)
        
        # Excel 列名配置
        excel_col_frame = ttk.Frame(excel_frame)
        excel_col_frame.pack(fill=tk.X)
        
        ttk.Label(excel_col_frame, text="IMEI 列名：", width=12).pack(side=tk.LEFT)
        self.excel_column_var = tk.StringVar(value="IMEI1")
        ttk.Entry(excel_col_frame, textvariable=self.excel_column_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Label(excel_col_frame, text="已加载：", width=8).pack(side=tk.LEFT)
        self.excel_count_var = tk.StringVar(value="0 条")
        ttk.Label(excel_col_frame, textvariable=self.excel_count_var).pack(side=tk.LEFT)
        
        # 打印机选择
        printer_frame = ttk.LabelFrame(tab, text="打印机", padding="10")
        printer_frame.pack(fill=tk.X, pady=(0, 10))
        
        printer_select_frame = ttk.Frame(printer_frame)
        printer_select_frame.pack(fill=tk.X)
        
        ttk.Label(printer_select_frame, text="选择打印机：", width=12).pack(side=tk.LEFT)
        self.printer_var = tk.StringVar()
        self.printer_combo = ttk.Combobox(printer_select_frame, textvariable=self.printer_var, state="readonly")
        self.printer_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(printer_select_frame, text="刷新", command=self.refresh_printers).pack(side=tk.RIGHT)
        
        # 打印配置
        config_frame = ttk.LabelFrame(tab, text="打印配置", padding="10")
        config_frame.pack(fill=tk.X, pady=(0, 10))
        
        config_row = ttk.Frame(config_frame)
        config_row.pack(fill=tk.X)
        
        ttk.Label(config_row, text="打印份数：", width=12).pack(side=tk.LEFT)
        self.copies_var = tk.IntVar(value=1)
        ttk.Spinbox(config_row, from_=1, to=100, textvariable=self.copies_var, width=10).pack(side=tk.LEFT, padx=(0, 20))
        
        self.verify_excel_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_row, text="打印前校验 Excel 数据", variable=self.verify_excel_var).pack(side=tk.LEFT)
        
        # 按钮区域
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(btn_frame, text="输入 IMEI 并打印", command=self.show_imei_input_dialog).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="批量导入 IMEI", command=self.import_imei_file).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="校验 Excel 数据", command=self.verify_excel_data).pack(side=tk.LEFT)
        
        # 状态显示
        status_frame = ttk.LabelFrame(tab, text="打印状态", padding="10")
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        self.print_status = tk.Text(status_frame, state=tk.DISABLED, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=self.print_status.yview)
        self.print_status.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.print_status.pack(fill=tk.BOTH, expand=True)
    
    def create_history_tab(self):
        """创建历史记录选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="历史记录")
        
        # 搜索框
        search_frame = ttk.Frame(tab)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="搜索 IMEI：").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.search_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 10))
        ttk.Button(search_frame, text="搜索", command=self.search_records).pack(side=tk.LEFT)
        ttk.Button(search_frame, text="显示全部", command=self.refresh_history).pack(side=tk.LEFT, padx=(5, 0))
        
        # 记录列表
        history_frame = ttk.LabelFrame(tab, text="打印记录", padding="10")
        history_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.history_tree = ttk.Treeview(history_frame, columns=("imei", "time", "copies", "status"), show="headings")
        self.history_tree.heading("imei", text="IMEI")
        self.history_tree.heading("time", text="打印时间")
        self.history_tree.heading("copies", text="份数")
        self.history_tree.heading("status", text="状态")
        self.history_tree.column("imei", width=200)
        self.history_tree.column("time", width=180)
        self.history_tree.column("copies", width=80)
        self.history_tree.column("status", width=80)
        self.history_tree.pack(fill=tk.BOTH, expand=True)
        
        # 按钮区域
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="导出 CSV", command=self.export_csv).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="导出 Excel", command=self.export_excel).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="清空记录", command=self.clear_records).pack(side=tk.LEFT)
        
        # 刷新历史
        self.refresh_history()
    
    def create_stats_tab(self):
        """创建统计选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="统计")
        
        # 统计卡片
        stats_frame = ttk.Frame(tab)
        stats_frame.pack(fill=tk.X, pady=(0, 20))
        
        # 今日打印
        today_card = ttk.LabelFrame(stats_frame, text="今日打印", padding="15")
        today_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.today_count_var = tk.StringVar(value="0")
        ttk.Label(today_card, textvariable=self.today_count_var, font=("微软雅黑", 24, "bold")).pack()
        ttk.Label(today_card, text="个 IMEI").pack()
        
        # 总打印
        total_card = ttk.LabelFrame(stats_frame, text="总打印", padding="15")
        total_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.total_count_var = tk.StringVar(value="0")
        ttk.Label(total_card, textvariable=self.total_count_var, font=("微软雅黑", 24, "bold")).pack()
        ttk.Label(total_card, text="个 IMEI").pack()
        
        # 今日份数
        copies_card = ttk.LabelFrame(stats_frame, text="今日份数", padding="15")
        copies_card.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.today_copies_var = tk.StringVar(value="0")
        ttk.Label(copies_card, textvariable=self.today_copies_var, font=("微软雅黑", 24, "bold")).pack()
        ttk.Label(copies_card, text="份").pack()
        
        # 最近打印记录
        recent_frame = ttk.LabelFrame(tab, text="最近打印", padding="10")
        recent_frame.pack(fill=tk.BOTH, expand=True)
        
        self.recent_tree = ttk.Treeview(recent_frame, columns=("imei", "time", "copies"), show="headings")
        self.recent_tree.heading("imei", text="IMEI")
        self.recent_tree.heading("time", text="时间")
        self.recent_tree.heading("copies", text="份数")
        self.recent_tree.pack(fill=tk.BOTH, expand=True)
        
        # 刷新统计
        self.refresh_stats()
    
    def init_bartender(self):
        """初始化 BarTender"""
        if not HAS_WIN32COM:
            self.update_status("警告：未安装 pywin32，BarTender 功能不可用")
            return
        
        try:
            # 初始化 COM
            pythoncom.CoInitialize()
            
            # 尝试连接 BarTender
            self.bt_app = win32com.client.Dispatch("BarTender.Application")
            self.bt_app.Visible = False
            self.update_status("BarTender 2022 R2 已连接")
            self.refresh_printers()
        except Exception as e:
            self.update_status(f"BarTender 连接失败: {e}")
            self.bt_app = None
    
    def refresh_printers(self):
        """刷新打印机列表"""
        printer_names = []
        
        try:
            if sys.platform == 'win32':
                # 方法1: 使用 win32print
                try:
                    import win32print
                    printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
                    printer_names = [p[2] for p in printers]
                except Exception:
                    pass
                
                # 方法2: 使用 PowerShell 获取打印机
                if not printer_names:
                    try:
                        import subprocess
                        result = subprocess.run(
                            ['powershell', '-Command', 'Get-Printer | Select-Object -ExpandProperty Name'],
                            capture_output=True, text=True, timeout=10
                        )
                        if result.returncode == 0:
                            printer_names = [p.strip() for p in result.stdout.split('\n') if p.strip()]
                    except Exception:
                        pass
                
                # 方法3: 使用 wmic 获取打印机
                if not printer_names:
                    try:
                        import subprocess
                        result = subprocess.run(
                            ['wmic', 'printer', 'get', 'name'],
                            capture_output=True, text=True, timeout=10
                        )
                        if result.returncode == 0:
                            lines = result.stdout.split('\n')
                            printer_names = [line.strip() for line in lines[1:] if line.strip()]
                    except Exception:
                        pass
        except Exception as e:
            self.update_status(f"获取打印机列表失败: {e}")
        
        # 更新打印机列表
        self.printer_combo['values'] = printer_names
        if printer_names:
            self.printer_combo.current(0)
            self.update_status(f"找到 {len(printer_names)} 台打印机")
        else:
            self.update_status("未找到打印机，请检查打印机连接")
    
    def browse_template(self):
        """浏览模板文件"""
        file_path = filedialog.askopenfilename(
            title="选择 BarTender 模板",
            filetypes=[("BarTender 文件", "*.btw"), ("所有文件", "*.*")]
        )
        if file_path:
            self.template_path_var.set(file_path)
            self.save_config()
    
    def browse_excel(self):
        """浏览 Excel 文件"""
        file_path = filedialog.askopenfilename(
            title="选择 IMEI 数据 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if file_path:
            self.excel_path_var.set(file_path)
            self.excel_file_path = file_path
            self.load_excel_data()
            self.save_config()
    
    def load_excel_data(self):
        """加载 Excel 数据"""
        file_path = self.excel_path_var.get()
        if not file_path or not os.path.exists(file_path):
            self.excel_data = []
            self.excel_count_var.set("0 条")
            return
        
        try:
            column_name = self.excel_column_var.get().strip()
            if not column_name:
                messagebox.showwarning("警告", "请输入 IMEI 列名")
                return
            
            if file_path.endswith('.csv'):
                # CSV 文件
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    self.excel_data = [row.get(column_name, '').strip() for row in reader if row.get(column_name, '').strip()]
            else:
                # Excel 文件
                if not HAS_OPENPYXL:
                    messagebox.showerror("错误", "未安装 openpyxl，无法读取 Excel 文件")
                    return
                
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                
                # 获取列索引
                header = [cell.value for cell in ws[1]]
                if column_name not in header:
                    messagebox.showerror("错误", f"Excel 中未找到列 '{column_name}'")
                    wb.close()
                    return
                
                col_idx = header.index(column_name)
                
                # 读取数据
                self.excel_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if col_idx < len(row) and row[col_idx]:
                        self.excel_data.append(str(row[col_idx]).strip())
                
                wb.close()
            
            self.excel_count_var.set(f"{len(self.excel_data)} 条")
            self.update_status(f"已加载 {len(self.excel_data)} 条 IMEI 数据")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载 Excel 失败: {e}")
            self.excel_data = []
            self.excel_count_var.set("0 条")
    
    def verify_excel_data(self):
        """校验 Excel 数据"""
        if not self.excel_data:
            messagebox.showwarning("警告", "请先选择并加载 Excel 文件")
            return
        
        messagebox.showinfo("校验结果", f"Excel 数据已加载\n\n文件: {os.path.basename(self.excel_file_path)}\n列名: {self.excel_column_var.get()}\n数据量: {len(self.excel_data)} 条")
    
    def is_imei_in_excel(self, imei):
        """检查 IMEI 是否在 Excel 数据中"""
        if not self.excel_data:
            return True
        return imei.strip() in self.excel_data
    
    def show_imei_input_dialog(self):
        """显示 IMEI 输入弹窗"""
        dialog = IMEIInputDialog(self.root, self)
        dialog.show()
    
    def import_imei_file(self):
        """从文件导入 IMEI"""
        file_path = filedialog.askopenfilename(
            title="选择 IMEI 文件",
            filetypes=[("文本文件", "*.txt"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                imei_list = [line.strip() for line in content.split('\n') if line.strip()]
                
                if imei_list:
                    self.process_imei_list(imei_list)
                else:
                    messagebox.showwarning("警告", "文件中未找到 IMEI 数据")
            except Exception as e:
                messagebox.showerror("错误", f"文件读取失败: {e}")
    
    def is_imei_printed(self, imei):
        """检查 IMEI 是否已打印"""
        return any(r.imei == imei for r in self.print_records)
    
    def print_single_imei(self, imei, template_path, printer, datasource_name):
        """打印单个 IMEI"""
        try:
            # 打开模板
            self.bt_format = self.bt_app.Formats.Open(template_path, False, "")
            
            # 设置数据源
            self.bt_format.SetNamedSubStringValue(datasource_name, imei)
            
            # 设置打印机
            self.bt_format.Printer = printer
            
            # 打印 - 忽略所有异常，只要执行到这里就视为成功
            try:
                self.bt_format.PrintOut(False, False)
            except:
                pass
            
            # 关闭模板
            self.bt_format.Close()
            
            # 打印成功
            return True
                
        except Exception as e:
            # 只有打开模板失败才算失败
            try:
                self.bt_format.Close()
            except:
                pass
            return False
    
    def process_imei_list(self, imei_list):
        """处理 IMEI 列表"""
        # 检查打印机
        printer = self.printer_var.get()
        if not printer:
            self.update_print_status("错误：请选择打印机", "error")
            return
        
        # 检查模板
        template_path = self.template_path_var.get()
        if not template_path or not os.path.exists(template_path):
            self.update_print_status("错误：请选择有效的 BarTender 模板文件", "error")
            return
        
        # 检查 BarTender
        if not self.bt_app:
            self.update_print_status("错误：BarTender 未连接，请检查安装", "error")
            return
        
        datasource_name = self.datasource_var.get()
        
        # 校验 Excel 数据
        if self.verify_excel_var.get() and self.excel_data:
            invalid_imei = [imei for imei in imei_list if not self.is_imei_in_excel(imei)]
            if invalid_imei:
                result = messagebox.askyesnocancel(
                    "数据不在文件中",
                    f"发现 {len(invalid_imei)} 个 IMEI 不在 Excel 数据文件中！\n\n"
                    f"无效 IMEI:\n{chr(10).join(invalid_imei[:5])}{'...' if len(invalid_imei) > 5 else ''}\n\n"
                    "点击「是」继续打印全部\n"
                    "点击「否」跳过无效 IMEI，只打印有效的\n"
                    "点击「取消」取消打印"
                )
                if result is None:  # 取消
                    return
                elif not result:  # 跳过无效
                    imei_list = [imei for imei in imei_list if self.is_imei_in_excel(imei)]
                    if not imei_list:
                        self.update_print_status("没有有效的 IMEI 可以打印", "error")
                        return
        
        # 检查已打印的 IMEI
        printed_imei = [imei for imei in imei_list if self.is_imei_printed(imei)]
        if printed_imei:
            result = messagebox.askyesnocancel(
                "数据重复",
                f"发现 {len(printed_imei)} 个 IMEI 已经打印过！\n\n"
                f"已打印的 IMEI:\n{chr(10).join(printed_imei[:5])}{'...' if len(printed_imei) > 5 else ''}\n\n"
                "点击「是」继续打印全部（包括已打印的）\n"
                "点击「否」跳过已打印的，只打印新的\n"
                "点击「取消」取消打印"
            )
            if result is None:  # 取消
                return
            elif not result:  # 跳过已打印
                imei_list = [imei for imei in imei_list if not self.is_imei_printed(imei)]
                if not imei_list:
                    self.update_print_status("所有 IMEI 都已打印过", "error")
                    return
        
        # 开始打印（使用线程避免阻塞UI）
        self.clear_print_status()
        self.update_print_status(f"开始打印 {len(imei_list)} 个 IMEI...", "info")
        
        # 在新线程中执行打印
        thread = threading.Thread(target=self._do_print, args=(imei_list, template_path, printer, datasource_name))
        thread.daemon = True
        thread.start()
    
    def _do_print(self, imei_list, template_path, printer, datasource_name):
        """在后台线程中执行打印"""
        success_count = 0
        fail_count = 0
        
        for imei in imei_list:
            success = self.print_single_imei(imei, template_path, printer, datasource_name)
            
            if success:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                record = PrintRecord(imei, now, 1, "PASS")
                self.print_records.append(record)
                success_count += 1
                self.root.after(0, lambda m=f"PASS {imei}": self.update_print_status(m, "success"))
            else:
                fail_count += 1
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                record = PrintRecord(imei, now, 1, "FAIL")
                self.print_records.append(record)
                self.root.after(0, lambda m=f"FAIL {imei}": self.update_print_status(m, "error"))
        
        # 保存配置和记录
        self.save_config()
        self.save_records()
        
        # 在主线程中刷新UI
        self.root.after(0, self.refresh_history)
        self.root.after(0, self.refresh_stats)
        self.root.after(0, lambda: self.update_print_status(f"\n打印完成：成功 {success_count}，失败 {fail_count}", "info"))
        self.root.after(0, lambda: self.update_status(f"打印完成：成功 {success_count}，失败 {fail_count}"))
    
    def clear_print_status(self):
        """清空打印状态"""
        self.print_status.config(state=tk.NORMAL)
        self.print_status.delete("1.0", tk.END)
        self.print_status.config(state=tk.DISABLED)
    
    def update_print_status(self, message, status="info"):
        """更新打印状态（带颜色）"""
        self.print_status.config(state=tk.NORMAL)
        
        # 配置颜色标签
        self.print_status.tag_configure("success", foreground="green")
        self.print_status.tag_configure("error", foreground="red")
        self.print_status.tag_configure("info", foreground="black")
        
        # 插入带颜色的文本
        tag = status if status in ["success", "error", "info"] else "info"
        self.print_status.insert(tk.END, message + "\n", tag)
        self.print_status.see(tk.END)
        self.print_status.config(state=tk.DISABLED)
        self.root.update_idletasks()
    
    def refresh_history(self):
        """刷新历史记录"""
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        for record in reversed(self.print_records):
            self.history_tree.insert('', 0, values=(
                record.imei,
                record.print_time,
                record.copies,
                record.status
            ))
    
    def search_records(self):
        """搜索记录"""
        keyword = self.search_var.get().strip()
        if not keyword:
            self.refresh_history()
            return
        
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        for record in reversed(self.print_records):
            if keyword.lower() in record.imei.lower():
                self.history_tree.insert('', 0, values=(
                    record.imei,
                    record.print_time,
                    record.copies,
                    record.status
                ))
    
    def refresh_stats(self):
        """刷新统计"""
        today = datetime.now().strftime("%Y-%m-%d")
        
        today_records = [r for r in self.print_records if r.print_time.startswith(today)]
        today_count = len(today_records)
        today_copies = sum(r.copies for r in today_records)
        total_count = len(self.print_records)
        
        self.today_count_var.set(str(today_count))
        self.total_count_var.set(str(total_count))
        self.today_copies_var.set(str(today_copies))
        
        for item in self.recent_tree.get_children():
            self.recent_tree.delete(item)
        
        for record in reversed(self.print_records[-20:]):
            self.recent_tree.insert('', 0, values=(
                record.imei,
                record.print_time,
                record.copies
            ))
    
    def export_csv(self):
        """导出 CSV"""
        file_path = filedialog.asksaveasfilename(
            title="导出 CSV",
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")]
        )
        
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(["IMEI", "打印时间", "份数", "状态"])
                    for record in self.print_records:
                        writer.writerow([record.imei, record.print_time, record.copies, record.status])
                
                messagebox.showinfo("成功", f"已导出 {len(self.print_records)} 条记录")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {e}")
    
    def export_excel(self):
        """导出 Excel"""
        if not HAS_OPENPYXL:
            messagebox.showerror("错误", "未安装 openpyxl，无法导出 Excel")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="导出 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "打印记录"
                
                ws.append(["IMEI", "打印时间", "份数", "状态"])
                for record in self.print_records:
                    ws.append([record.imei, record.print_time, record.copies, record.status])
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"已导出 {len(self.print_records)} 条记录")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {e}")
    
    def clear_records(self):
        """清空记录"""
        if messagebox.askyesno("确认", "确定要清空所有打印记录吗？"):
            self.print_records.clear()
            self.save_records()
            self.refresh_history()
            self.refresh_stats()
            messagebox.showinfo("成功", "记录已清空")
    
    def open_settings(self):
        """打开设置"""
        SettingsWindow(self.root, self)
    
    def load_config(self):
        """加载配置"""
        missing_files = []
        
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                    # 加载模板路径
                    template_path = config.get('template_path', '')
                    if template_path and not os.path.exists(template_path):
                        missing_files.append(f"BarTender 模板: {os.path.basename(template_path)}")
                    self.template_path_var.set(template_path)
                    
                    # 加载数据源名称
                    self.datasource_var.set(config.get('datasource', 'IMEI1'))
                    
                    # 加载 Excel 路径
                    excel_path = config.get('excel_path', '')
                    if excel_path and not os.path.exists(excel_path):
                        missing_files.append(f"Excel 文件: {os.path.basename(excel_path)}")
                    self.excel_path_var.set(excel_path)
                    self.excel_file_path = excel_path
                    
                    # 加载 Excel 列名
                    self.excel_column_var.set(config.get('excel_column', 'IMEI1'))
                    
                    # 加载打印机
                    self.printer_var.set(config.get('printer', ''))
                    
                    # 加载校验选项
                    self.verify_excel_var.set(config.get('verify_excel', True))
                    
        except Exception as e:
            print(f"加载配置失败: {e}")
        
        return missing_files
    
    def save_config(self):
        """保存配置"""
        try:
            config = {
                'template_path': self.template_path_var.get(),
                'datasource': self.datasource_var.get(),
                'excel_path': self.excel_path_var.get(),
                'excel_column': self.excel_column_var.get(),
                'printer': self.printer_var.get(),
                'verify_excel': self.verify_excel_var.get()
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存配置失败: {e}")
    
    def load_records(self):
        """加载打印记录"""
        try:
            if os.path.exists(self.records_file):
                with open(self.records_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        record = PrintRecord(
                            imei=row.get('imei', ''),
                            print_time=row.get('print_time', ''),
                            copies=int(row.get('copies', 1)),
                            status=row.get('status', '成功')
                        )
                        self.print_records.append(record)
        except Exception:
            pass
    
    def save_records(self):
        """保存打印记录"""
        try:
            with open(self.records_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['imei', 'print_time', 'copies', 'status'])
                for record in self.print_records:
                    # 确保IMEI作为字符串保存，添加前缀防止Excel截断
                    imei_str = str(record.imei) if record.imei else ''
                    writer.writerow([imei_str, record.print_time, record.copies, record.status])
        except Exception as e:
            print(f"保存记录失败: {e}")
    
    def load_records(self):
        """加载打印记录"""
        try:
            if os.path.exists(self.records_file):
                with open(self.records_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)  # 跳过表头
                    for row in reader:
                        if len(row) >= 4:
                            record = PrintRecord(
                                imei=row[0].strip(),
                                print_time=row[1].strip(),
                                copies=int(row[2]) if row[2].isdigit() else 1,
                                status=row[3].strip()
                            )
                            self.print_records.append(record)
        except Exception as e:
            print(f"加载记录失败: {e}")
        
        # 清理 COM
        if HAS_WIN32COM:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        
        self.root.destroy()
    
    def run(self):
        """运行应用"""
        self.root.mainloop()


class IMEIInputDialog:
    """IMEI 输入弹窗"""
    
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        
    def show(self):
        """显示弹窗"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("输入 IMEI")
        dialog.geometry("400x200")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # 主框架
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 说明
        ttk.Label(main_frame, text="输入要打印的 IMEI（每行一个）：", font=("微软雅黑", 11)).pack(anchor=tk.W, pady=(0, 10))
        
        # IMEI 输入框
        imei_text = tk.Text(main_frame, wrap=tk.WORD, height=3, width=40)
        imei_text.pack(fill=tk.X, pady=(0, 10))
        imei_text.focus_set()
        
        # 监听内容变化，自动调整高度
        def on_content_change(event=None):
            content = imei_text.get("1.0", tk.END)
            lines = content.count('\n') + 1
            new_height = min(max(lines, 3), 10)
            imei_text.config(height=new_height)
        
        imei_text.bind('<KeyRelease>', on_content_change)
        
        # 按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        def on_print():
            imei_text_content = imei_text.get("1.0", tk.END).strip()
            if not imei_text_content:
                return
            
            imei_list = [line.strip() for line in imei_text_content.split('\n') if line.strip()]
            dialog.destroy()
            self.app.process_imei_list(imei_list)
        
        def on_enter(event):
            content = imei_text.get("1.0", tk.END).strip()
            if content:
                lines = [l.strip() for l in content.split('\n') if l.strip()]
                if len(lines) == 1:
                    on_print()
        
        imei_text.bind('<Return>', on_enter)
        
        ttk.Button(btn_frame, text="打印", command=on_print).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)


class SettingsWindow:
    """设置窗口"""
    
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        
        self.window = tk.Toplevel(parent)
        self.window.title("设置")
        self.window.geometry("400x300")
        self.window.transient(parent)
        self.window.grab_set()
        
        self.create_ui()
    
    def create_ui(self):
        """创建界面"""
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="设置", font=("微软雅黑", 12, "bold")).pack(anchor=tk.W, pady=(0, 15))
        
        # 默认打印机
        ttk.Label(main_frame, text="默认打印机：").pack(anchor=tk.W)
        self.default_printer_var = tk.StringVar(value=self.app.printer_var.get())
        ttk.Entry(main_frame, textvariable=self.default_printer_var).pack(fill=tk.X, pady=(0, 10))
        
        # 自动校验
        self.auto_verify_var = tk.BooleanVar(value=self.app.verify_excel_var.get())
        ttk.Checkbutton(main_frame, text="打印前自动校验 Excel 数据", variable=self.auto_verify_var).pack(anchor=tk.W, pady=(0, 10))
        
        # 数据源名称
        ttk.Label(main_frame, text="BarTender 数据源名称：").pack(anchor=tk.W)
        self.datasource_var = tk.StringVar(value=self.app.datasource_var.get())
        ttk.Entry(main_frame, textvariable=self.datasource_var).pack(fill=tk.X, pady=(0, 20))
        
        # 按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="保存", command=self.save).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(btn_frame, text="取消", command=self.window.destroy).pack(side=tk.RIGHT)
    
    def save(self):
        """保存设置"""
        self.app.printer_var.set(self.default_printer_var.get())
        self.app.verify_excel_var.set(self.auto_verify_var.get())
        self.app.datasource_var.set(self.datasource_var.get())
        self.app.save_config()
        self.window.destroy()


def main():
    """主函数"""
    app = BarTenderPrintApp()
    app.run()


if __name__ == "__main__":
    main()
